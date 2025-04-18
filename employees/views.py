from django.shortcuts import render,redirect, get_object_or_404
from .models import employee_model
from django.contrib.auth import authenticate, login
from app1.models import department as department_model, team as team_model
from app1.forms import department_form
from app1.forms import employee_form
from app1.forms import LoginForm
import openpyxl
from openpyxl import Workbook
from django.contrib.auth.models import User
from django.contrib.auth.views import LogoutView
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.contrib.auth.decorators import login_required
from datetime import date
from django.http import HttpResponse, HttpResponseForbidden
from datetime import datetime

# Create your views here.
def get_employees(request,username,id):
    user =get_object_or_404(User,username=username)
    if request.user != user :
        return HttpResponse('You have not permission to access this page')
    else:
        employee=employee_model.objects.get(employee_id=username)
        employee_list=employee_model.objects.filter(department_id=id)
        department=department_model.objects.get(department_id=id)
        team_list=team_model.objects.filter(department_id=id)
    return render(request,'employees.html',{'employee_list':employee_list,'department':department, 'username': username, 'team_list': team_list, 'employee': employee})


@login_required     
def upload_excel_emp(request, username):
    user =get_object_or_404(User,username=username)    
    if request.user != user or not user.is_superuser :
        return HttpResponse('You have not permission to access this page')
    else:     
        if request.method == 'POST' and request.FILES.get('excel_file'):
            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            upload_file_url = fs.url(filename)
            wb = openpyxl.load_workbook(fs.path(filename))
            sheet = wb.active
            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                employee_id, name, leader,team_id,team_leader,department_id,manager_dep, grade, email=row
                if employee_id is None:
                    messages.error(request, f"Dòng {index}: Thiếu employee_id.")
                    continue
                employee_id = str(employee_id).strip()
                name = str(name).strip()
                leader = str(leader).strip()
                team_id = str(team_id).strip()
                team_leader = str(team_leader).strip()
                department_id = str(department_id).strip()
                manager_dep = str(manager_dep).strip()
                grade = str(grade).strip()                
                email = str(email).strip() if email else ""              

                if not employee_id:
                    messages.error(request, f"Dòng {index}: employee_id rỗng.")
                    continue

                if not employee_model.objects.filter(employee_id=employee_id).exists():
                    try:
                        employee_model.objects.create(
                            employee_id=employee_id,
                            name=name,
                            leader=leader,
                            team_id=team_id,
                            team_leader=team_leader,
                            department_id=department_id,
                            manager_dep=manager_dep,
                            grade=grade,
                            email=email,
                            Created_date=date.today()
                        )
                    except Exception as e:
                        messages.error(request, f"Dòng {index}: Lỗi tạo employee {employee_id} - {e}")
                else:
                    messages.warning(request, f"Dòng {index}: employee_id {employee_id} đã tồn tại.")

            return redirect('upload_excel_emp', username=username)  # ✅ đảm bảo trả về sau khi xử lý
        # return HttpResponse('Finish upload user list')
    # ✅ đảm bảo có return nếu không phải POST hoặc không có file
        return render(request, 'upload_excel_emp.html')
    
def export_employees_to_excel(request):
    # Tạo Workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    # Thêm header
    headers = ['employee_id', 'name', 'leader', 'email','grade','team_id','team_leader','department_id','manager_dep']
    sheet.append(headers)

    # Lấy dữ liệu từ database
    employees = employee_model.objects.all().values_list('employee_id', 'name', 'leader', 'email','grade','team_id','team_leader','department_id','manager_dep')
    for employee in employees:
        sheet.append(employee)

    # Tạo response với file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    workbook.save(response)

    return response